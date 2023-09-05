import datetime
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import time
import sys, os

#Returns pool prices for the last 24 hours as a 2d array where
#hourlyPrices[X][0] = date/hour & hourlyPrices[X][1] = price
def getPoolPrice():
    # Send a get request to get website data and convert to string
    response = requests.get('http://ets.aeso.ca/ets_web/ip/Market/Reports/SMPriceReportServlet')
    websiteText = response.text

    #Parse the HTML for the correct table, find all entries in correct table
    soup = BeautifulSoup(websiteText, 'html.parser')
    table = soup.find_all('table')[2]
    rows = table.find_all('tr')

    #For storing updates in
    hourlyPrices = []

    # For each row in the table
    for row in rows:
        cols = row.find_all('td')
        # If not the header row, extract data and append
        if len(cols) >= 2:
            hour = cols[0].text.strip()
            price = cols[1].text.strip()
            hourlyPrices.append((hour, price))

    return hourlyPrices

#Returns TNG number as a float
def getTNG():
    #Grab the website data, filter it down to just show the H.R. Milner row as array
    response = requests.get('http://ets.aeso.ca/ets_web/ip/Market/Reports/CSDReportServlet')
    soup = BeautifulSoup(response.text, 'html.parser')
    table = soup.find_all('table')[9]
    hrmTable = table.find_all('tr')[14].find_all('td')
    #Return TNG for the hour as a float
    return ((hrmTable[2].text.strip()))

#Gets relevant data and update Excel spreadsheet as needed
def updateSpreadSheet(excelPath):
    #Grab values from other functions
    prices = getPoolPrice()
    tng = getTNG()

    #Open Excel spreadsheet
    wb = load_workbook(excelPath)
    sheet = wb.active

    #Transform data/existing spreadsheet data make it easier to manipulate via dictionaries
    priceDict = {hour: (float(price) if price != '-' else float(-1))
                    for hour, price in prices}
    existingData = {str(row[0]): str(row[1]) for row in
                    sheet.iter_rows(min_row=2, max_col=2, values_only=True)}

    #Add newest hour only if not added previously
    newestPrice = list(priceDict.items())[0]
    prevPrice = list(priceDict.items())[1]
    if newestPrice[0] not in existingData:
        sheet.append([newestPrice[0], newestPrice[1], '-'])
        #Update TNG for last hour once new hour added if for last hour
        if (int(prevPrice[0][11:]) == int(newestPrice[0][11:])-1) and (prevPrice[0][0:10] == newestPrice[0][0:10]):
            sheet.cell(row = sheet.max_row-1, column = 3).value = tng
        elif isYesterday(newestPrice[0][0:10], prevPrice[0][0:10]) \
                and prevPrice[0][11:] == '24' and newestPrice[0][11:] == '01':
            sheet.cell(row=sheet.max_row - 1, column=3).value = tng


    #Update previous values
    if (sheet.max_row - 24) <= 1:
        startRow = 2
    else:
        startRow = sheet.max_row - 24
    for row in sheet.iter_rows(min_row=startRow, max_row=sheet.max_row):
        try:
            #If price for hour in sheet is null & there is an actual price, update it
            if row[1].value == -1 and priceDict[row[0].value] != -1:
                row[1].value = priceDict[row[0].value]
                print('Found price for', row[0].value)
        except KeyError: #Nothing, just move on
            print('Unable to find the price for', row[0].value)
        finally:  #Here to ensure loop continues
            x=1
    wb.save(excelPath)

#Add the price from last hour (using just for initial setup)
def addLastFullHour(excelPath):
    # Grab values from other function
    prices = getPoolPrice()
    wb = load_workbook(excelPath)
    sheet = wb.active
    priceDict = {hour: (float(price) if price != '-' else float(-1)) for hour, price in prices}
    secondNewestPrice = list(priceDict.items())[1]
    sheet.append([secondNewestPrice[0], secondNewestPrice[1], '-'])
    # Save the new data
    wb.save(excelPath)

#Check if two strings represent sequqntial days
# (dayOne = today?, dayTwo = day before?)
def isYesterday(dayOne, dayTwo):
    oneArr = dayOne.split('/')
    twoArr = dayTwo.split('/')
    #print()
    #If years equal
    if oneArr[2] == twoArr[2]:
        #If months equal
        if oneArr[0] == twoArr[0]:
            #If previous day
            if int(oneArr[1]) == (int(twoArr[1]) + 1):
                return True
        #If month turnover
        elif (int(oneArr[0]) == (int(twoArr[0]) + 1)) and (oneArr[1] == '01'):
            #Feb
            if (int(twoArr[0]) == 2) and ((int(twoArr[1]) == 28) or (int(twoArr[1]) == 29)):
                return True
            #30 day months
            elif (int(twoArr[1]) == 30) and ((int(twoArr[0]) == 4) or (int(twoArr[0]) == 6)
                        or (int(twoArr[0]) == 9) or (int(twoArr[0]) == 11)):
                return True
            #31 day months
            elif (int(twoArr[1]) == 31) and ((int(twoArr[0]) == 1) or (int(twoArr[0]) == 3)
                        or (int(twoArr[0]) == 5) or int((twoArr[0]) == 7) or (int(twoArr[0]) == 8)
                        or (int(twoArr[0]) == 10) or int((twoArr[0]) == 12)):
                return True
    #If new years
    elif((int(oneArr[2]) == int(twoArr[2])+1) and (int(oneArr[1]) == 1) and
        (int(oneArr[0]) == 1) and (int(twoArr[1]) == 31) and (int(twoArr[0]) == 12)):
        return True
    return False

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    #Get directory
    if getattr(sys, 'frozen', False): #EXE
        dirPath = os.path.dirname(sys.executable)
    else:   #Python script
        dirPath = os.path.dirname(os.path.abspath(__file__))

    #Load workbook from directory
    excelPath = str(dirPath) + '/gasPrices.xlsx'
    wb = load_workbook(excelPath)
    sheet = wb.active
    if(sheet.max_row == 1):
        addLastFullHour(excelPath)
    while True:
        #Update spreadsheet and print confirmation
        updateSpreadSheet(excelPath)
        print(datetime.datetime.now().strftime("%m-%d-%Y %H:%M") ,"Sheet updated, waiting one hour...")
        # Wait for 1 hour
        time.sleep(3600)